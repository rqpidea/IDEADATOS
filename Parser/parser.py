from openpyxl import load_workbook 
import csv, sys
datos = sys.argv[1] + '.xlsx'
razonsocial = 'r_social.csv'
output = sys.argv[1] + '.csv'

concepto = ['PROMOCION DE OBRAS Y ACTIVIDADES MPALES',
			'PRODUCCION DE SPOTS DE PUBLICIDAD']
r_social = set()


r_file = open(razonsocial, 'rb')
csv_f = csv.reader(r_file, delimiter=',', quotechar='|')
print 'Leyendo datos de proveedores'
for row in csv_f:
  r_social.add(''.join(row))
r_file.close()

wb = load_workbook(datos, use_iterators=True)
sheet = wb.worksheets[0]
ws = wb.active
row_count = sheet.get_highest_row() -1
r_file = open(razonsocial, 'w+')
r_writer = csv.writer(r_file, delimiter=',')
print 'Buscando proveedores nuevos'
for row_idx in range(1, row_count):
	if(ws.cell('E%s'%(row_idx)).value in concepto):
		r_social.add(ws.cell('C%s'%(row_idx)).value)
print 'Actualizando proveedores'
for social_idx in r_social:
	r_writer.writerow([social_idx])
r_file.close()

o_file = open(output, 'w+')
o_writer = csv.writer(o_file, delimiter=',')
o_writer.writerow(['Beneficiario', 'Concepto', 'Importe'])
print 'Creando archivo'
for row_idx in range(1, row_count):
	if(ws.cell('C%s'%(row_idx)).value in r_social):
		o_writer.writerow([ws.cell('C%s'%(row_idx)).value, ws.cell('E%s'%(row_idx)).value, ws.cell('F%s'%(row_idx)).value])

o_file.close()
print 'Termino'